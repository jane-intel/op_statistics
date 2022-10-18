from collections import defaultdict
from pathlib import Path
from typing import Dict
import logging as log
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter


def apply_filter(sheet):
    sheet.auto_filter.ref = 'A1:' + get_column_letter(sheet.max_column) + str(sheet.max_row)


def data_to_sheet(name: str, wb: Workbook, data: pd.DataFrame):
    sheet = wb.create_sheet(name)
    sheet.append([""] + list(data.columns))
    values = data.values.tolist()
    for i, item in enumerate(data.index):
        sheet.append([item] + list(map(lambda x: "" if pd.isna(x) else int(x), values[i])))
    apply_filter(sheet)


def dict_to_sheet(name: str, wb: Workbook, data: Dict, head: list):
    sheet = wb.create_sheet(name)
    sheet.append(head)
    for name, num in data.items():
        sheet.append([name, num])
    apply_filter(sheet)


def report(ops_per_model: dict, summary: defaultdict, errors: dict):
    wb = Workbook()

    dict_to_sheet("Summary", wb, summary, ["Op name", "Num ops"])

    data_frame = pd.DataFrame.from_dict(ops_per_model, "index")
    data_to_sheet("By model", wb, data_frame)

    data_frame = pd.DataFrame.from_dict(ops_per_model)
    data_to_sheet("By op", wb, data_frame)

    dict_to_sheet("Errors", wb, errors, ["Model", "Error"])

    wb.remove(wb.active)
    save_excel_path = Path().cwd() / "report.xlsx"
    wb.save(save_excel_path)
    log.info("[ SUCCESS ] Excel file saved: {}".format(save_excel_path))


def all_files_by_ext(directory: str, ext: str):
    return sorted([str(f) for f in list(Path(directory).rglob("*.{}".format(ext)))])
